"""
Staff bulk import integration tests.
Run inside Docker: docker compose exec api pytest app/tests/test_staff_import.py -v

Fixtures (client, auth) come from conftest.py.
"""
import io
import pytest
from httpx import AsyncClient

from app.models.school import School
from app.services.staff_import_constants import DATA_START, _COLS, make_sentinel


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_xlsx(rows: list[list], school_code: str) -> bytes:
    """Build a minimal .xlsx with the correct school-specific sentinel and data rows."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Staff Data"
    ws["N1"] = make_sentinel(school_code)
    for col, label, *_ in _COLS:
        ws[f"{col}3"] = label
    for r_idx, row_vals in enumerate(rows):
        row_num = DATA_START + r_idx
        for c_idx, val in enumerate(row_vals):
            ws.cell(row=row_num, column=c_idx + 1, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_VALID_ROW = ["IMP001", "Ama", None, "Boateng", "FEMALE", None, None, None, None, None, None, None]


# ── Template download ─────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_download_template_returns_xlsx(client: AsyncClient, auth: dict):
    resp = await client.get("/staff/import/template", headers=auth)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(_XLSX_CT)
    assert len(resp.content) > 5_000   # non-trivial workbook size


@pytest.mark.asyncio
async def test_template_contains_sentinel(client: AsyncClient, auth: dict, school: School):
    """The downloaded template must contain the school-specific sentinel."""
    from openpyxl import load_workbook
    resp = await client.get("/staff/import/template", headers=auth)
    wb = load_workbook(io.BytesIO(resp.content), data_only=True)
    ws = wb["Staff Data"]
    assert ws["N1"].value == make_sentinel(school.school_code)


@pytest.mark.asyncio
async def test_template_has_all_column_headers(client: AsyncClient, auth: dict):
    from openpyxl import load_workbook
    resp = await client.get("/staff/import/template", headers=auth)
    wb = load_workbook(io.BytesIO(resp.content), data_only=True)
    ws = wb["Staff Data"]
    headers = [ws[f"{col}3"].value for col, *_ in _COLS]
    assert all(h is not None for h in headers)


# ── Upload processing ─────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_import_all_valid(client: AsyncClient, auth: dict, school: School):
    rows = [
        ["IMP001", "Ama",  None, "Boateng", "FEMALE", None, None, None, None, None, None, None],
        ["IMP002", "Kofi", None, "Asante",  "MALE",   None, None, None, None, None, None, None],
    ]
    resp = await client.post(
        "/staff/import",
        files={"file": ("staff.xlsx", _make_xlsx(rows, school.school_code), _XLSX_CT)},
        headers=auth,
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["created"] == 2
    assert data["failed"] == 0
    assert data["errors"] == []
    assert "batch_id" in data


@pytest.mark.asyncio
async def test_import_partial_success(client: AsyncClient, auth: dict, school: School):
    """Two valid rows + one with missing required last_name → 2 created, 1 failed."""
    rows = [
        ["IMP001", "Ama",  None, "Boateng", None, None, None, None, None, None, None, None],
        ["IMP002", "Kofi", None, "Asante",  None, None, None, None, None, None, None, None],
        ["IMP003", "Yaw",  None, None,      None, None, None, None, None, None, None, None],
    ]
    resp = await client.post(
        "/staff/import",
        files={"file": ("staff.xlsx", _make_xlsx(rows, school.school_code), _XLSX_CT)},
        headers=auth,
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["created"] == 2
    assert data["failed"] == 1
    assert data["total_rows"] == 3
    assert len(data["errors"]) == 1
    assert data["errors"][0]["row"] == DATA_START + 2
    assert data["errors"][0]["ref"] == "IMP003"


@pytest.mark.asyncio
async def test_import_duplicate_staff_number(client: AsyncClient, auth: dict, school: School):
    """Duplicate staff_number in the same upload → first created, second failed."""
    rows = [
        ["DUP001", "Abena", None, "Mensah", None, None, None, None, None, None, None, None],
        ["DUP001", "Akua",  None, "Asante", None, None, None, None, None, None, None, None],
    ]
    resp = await client.post(
        "/staff/import",
        files={"file": ("staff.xlsx", _make_xlsx(rows, school.school_code), _XLSX_CT)},
        headers=auth,
    )
    data = resp.json()
    assert data["created"] == 1
    assert data["failed"] == 1
    assert "already exists" in data["errors"][0]["error"]


@pytest.mark.asyncio
async def test_import_rejects_existing_staff_number(client: AsyncClient, auth: dict, school: School):
    """Staff number already in DB is rejected even in a fresh upload."""
    await client.post("/staff", json={"staff_number": "PRE001", "first_name": "Existing", "last_name": "Staff"}, headers=auth)
    rows = [["PRE001", "New", None, "Person", None, None, None, None, None, None, None, None]]
    resp = await client.post(
        "/staff/import",
        files={"file": ("staff.xlsx", _make_xlsx(rows, school.school_code), _XLSX_CT)},
        headers=auth,
    )
    data = resp.json()
    assert data["created"] == 0
    assert data["failed"] == 1


@pytest.mark.asyncio
async def test_import_cross_school_sentinel_rejected(client: AsyncClient, auth: dict):
    """A template from a different school (wrong sentinel) must return 422 with a helpful message."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Staff Data"
    ws["N1"] = make_sentinel("OTHER_SCHOOL")   # different school code
    buf = io.BytesIO()
    wb.save(buf)
    resp = await client.post(
        "/staff/import",
        files={"file": ("other.xlsx", buf.getvalue(), _XLSX_CT)},
        headers=auth,
    )
    assert resp.status_code == 422
    assert "OTHER_SCHOOL" in resp.json()["detail"]


@pytest.mark.asyncio
async def test_import_wrong_template_rejected(client: AsyncClient, auth: dict):
    """A generic .xlsx with no sentinel must return 422."""
    from openpyxl import Workbook
    wb = Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    resp = await client.post(
        "/staff/import",
        files={"file": ("generic.xlsx", buf.getvalue(), _XLSX_CT)},
        headers=auth,
    )
    assert resp.status_code == 422


@pytest.mark.asyncio
async def test_import_non_xlsx_rejected(client: AsyncClient, auth: dict):
    """Uploading a .csv should return 422 immediately."""
    resp = await client.post(
        "/staff/import",
        files={"file": ("staff.csv", b"staff_number,first_name\nTST001,Kwame", "text/csv")},
        headers=auth,
    )
    assert resp.status_code == 422
