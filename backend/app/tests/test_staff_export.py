"""
Staff register export tests — one unified endpoint (/staff/export/custom)
covering CSV, Excel, and PDF, all with caller-selected fields.
Run inside Docker: docker compose exec api pytest app/tests/test_staff_export.py -v
"""
import csv
import io

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.auth import StaffPosition
from app.models.staff import StaffCategory, StaffType
from app.models.school import School


def _staff(num: str = "TST001", **kw) -> dict:
    return {"staff_number": num, "first_name": "Kwame", "last_name": "Mensah", **kw}


@pytest.mark.asyncio
async def test_export_excel_basic(client: AsyncClient, auth: dict):
    await client.post("/staff", json=_staff(), headers=auth)
    resp = await client.get("/staff/export/custom?fields=staff_number,full_name&fmt=excel", headers=auth)
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    assert rows[1][header.index("Staff No.")] == "TST001"


@pytest.mark.asyncio
async def test_export_pdf_basic(client: AsyncClient, auth: dict):
    await client.post("/staff", json=_staff(), headers=auth)
    resp = await client.get("/staff/export/custom?fields=staff_number,full_name&fmt=pdf", headers=auth)
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content.startswith(b"%PDF")


@pytest.mark.asyncio
async def test_custom_export_csv_basic(client: AsyncClient, auth: dict):
    await client.post("/staff", json=_staff(), headers=auth)
    resp = await client.get("/staff/export/custom?fields=staff_number,full_name&fmt=csv", headers=auth)
    assert resp.status_code == 200
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8-sig"))))
    header, data_row = rows[0], rows[1]
    assert data_row[header.index("Staff No.")] == "TST001"


@pytest.mark.asyncio
async def test_export_rejects_unknown_format(client: AsyncClient, auth: dict):
    resp = await client.get("/staff/export/custom?fields=staff_number&fmt=doc", headers=auth)
    assert resp.status_code == 422


# ── Search parity with the on-screen list (services/staff_query.py) ─────────
# list_staff() matches a search term against name/staff-number AND
# category/position name; the export path used to only check name/staff
# number, so filtering the on-screen list by a position or category name
# would show matches on screen but silently drop them from the export.

@pytest.mark.asyncio
async def test_export_excel_search_matches_position_name(client: AsyncClient, auth: dict, db_session: AsyncSession):
    pos = await db_session.scalar(select(StaffPosition).where(StaffPosition.code == "BURSAR"))
    assert pos is not None, "Run seed_reference_data.py first"
    staff_id = (await client.post("/staff", json=_staff("BUR001"), headers=auth)).json()["id"]
    await client.patch(f"/staff/{staff_id}", json={"position_ids": [str(pos.id)]}, headers=auth)
    await client.post("/staff", json=_staff("OTH001", first_name="Ama", last_name="Boateng"), headers=auth)

    list_resp = await client.get("/staff?search=Bursar", headers=auth)
    assert [s["staff_number"] for s in list_resp.json()] == ["BUR001"]

    resp = await client.get("/staff/export/custom?fields=staff_number&fmt=excel&search=Bursar", headers=auth)
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    staff_numbers = [r[header.index("Staff No.")] for r in rows[1:]]
    assert staff_numbers == ["BUR001"], "export must match exactly what the on-screen search matched"


@pytest.mark.asyncio
async def test_custom_export_search_matches_category_name(
    client: AsyncClient, auth: dict, db_session: AsyncSession, school: School,
):
    cat = StaffCategory(
        school_id=school.id, name="Groundskeeping Staff", code="GROUNDS",
        staff_type=StaffType.NON_TEACHING, is_template=False, is_active=True,
    )
    db_session.add(cat)
    await db_session.flush()
    staff_id = (await client.post("/staff", json=_staff("GRD001"), headers=auth)).json()["id"]
    await client.patch(f"/staff/{staff_id}", json={"category_id": str(cat.id)}, headers=auth)
    await client.post("/staff", json=_staff("OTH002", first_name="Ama", last_name="Boateng"), headers=auth)

    list_resp = await client.get("/staff?search=Groundskeeping", headers=auth)
    assert [s["staff_number"] for s in list_resp.json()] == ["GRD001"]

    resp = await client.get(
        "/staff/export/custom?fields=staff_number&fmt=csv&search=Groundskeeping", headers=auth,
    )
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8-sig"))))
    header, data_rows = rows[0], rows[1:]
    staff_numbers = [r[header.index("Staff No.")] for r in data_rows]
    assert staff_numbers == ["GRD001"], "export must match exactly what the on-screen search matched"
