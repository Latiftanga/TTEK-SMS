"""
Student register export tests (CSV full export + custom field export).
Run inside Docker: docker compose exec api pytest app/tests/test_student_export.py -v
"""
import csv
import io
import uuid
from datetime import date, datetime

import pytest
from httpx import AsyncClient
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.auth import hash_password
from app.models.academic import AcademicYear, Class, ClassTeacher
from app.models.auth import LoginType, StaffPosition, User
from app.models.documents import GraduationRecord, GraduationType
from app.models.school import School
from app.models.students import StudentClassAssignment


def _student(num: str = "ADM001", **kw) -> dict:
    return {"admission_number": num, "first_name": "Ama", "last_name": "Boateng", **kw}


async def _login_as_position(
    client: AsyncClient, auth: dict, db_session: AsyncSession, school: School, position_code: str,
) -> tuple[dict, str]:
    """Create a staff member holding `position_code`, give them a login, and
    return (their bearer-token auth headers, their staff_member id) — mirrors
    test_students.py's helper."""
    pos = await db_session.scalar(select(StaffPosition).where(StaffPosition.code == position_code))
    assert pos is not None, "Run seed_reference_data.py first"

    staff_id = (await client.post("/staff", json={
        "staff_number": f"TST-{position_code}", "first_name": "Test", "last_name": position_code.title(),
    }, headers=auth)).json()["id"]
    await client.patch(f"/staff/{staff_id}", json={"position_ids": [str(pos.id)]}, headers=auth)

    email = f"{position_code.lower()}@presec-test.edu.gh"
    db_session.add(User(
        school_id=school.id, login_type=LoginType.EMAIL, email=email,
        password_hash=hash_password("Whatever123!"), is_active=True, staff_member_id=staff_id,
    ))
    await db_session.flush()

    resp = await client.post("/auth/login", json={
        "login_type": "EMAIL", "identifier": email, "password": "Whatever123!",
        "school_code": school.school_code,
    })
    assert resp.status_code == 200, resp.text
    return {"Authorization": f"Bearer {resp.json()['access_token']}"}, staff_id


async def _promote_student(
    client: AsyncClient, auth: dict, db_session: AsyncSession, school: School,
    school_class: Class, academic_year: AcademicYear,
) -> str:
    """Create a student with 2 concurrent is_active StudentClassAssignment rows —
    the old class (school_class, year 1) and a newer one (next_class, year 2) —
    mirroring a promoted-but-not-graduated student. Returns the student id."""
    sid = (await client.post("/students", json=_student("ADM001"), headers=auth)).json()["id"]

    next_year = AcademicYear(
        school_id=school.id, name="2025/2026",
        start_date=date(2025, 9, 1), end_date=date(2026, 7, 31), is_current=False,
    )
    db_session.add(next_year)
    await db_session.flush()
    next_class = Class(school_id=school.id, level="SHS", year_group=3, stream="A", is_active=True)
    db_session.add(next_class)
    await db_session.flush()

    db_session.add(StudentClassAssignment(
        school_id=school.id, student_id=sid, class_id=school_class.id,
        academic_year_id=academic_year.id, is_active=True,
    ))
    await db_session.flush()
    db_session.add(StudentClassAssignment(
        school_id=school.id, student_id=sid, class_id=next_class.id,
        academic_year_id=next_year.id, is_active=True,
    ))
    await db_session.flush()
    return sid


@pytest.mark.asyncio
async def test_export_csv_shows_current_class_for_promoted_student(
    client: AsyncClient, auth: dict, db_session: AsyncSession, school: School,
    school_class: Class, academic_year: AcademicYear,
):
    """A promoted (non-graduated) student has 2 active class assignments —
    the export must show the current (most recent) class, not the old one."""
    await _promote_student(client, auth, db_session, school, school_class, academic_year)

    resp = await client.get("/students/export/custom?fields=admission_number,current_class&fmt=csv", headers=auth)
    assert resp.status_code == 200
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8-sig"))))
    header, data_row = rows[0], rows[1]
    # SHS classes drop the redundant "SHS" word (school-wide format, matches the
    # class list / dashboards) — see student_display.py::_class_display_name.
    assert data_row[header.index("Current Class")] == "3 A"


@pytest.mark.asyncio
async def test_custom_export_shows_current_class_for_promoted_student(
    client: AsyncClient, auth: dict, db_session: AsyncSession, school: School,
    school_class: Class, academic_year: AcademicYear,
):
    """Same promoted-student scenario, exercised through the custom-field export."""
    await _promote_student(client, auth, db_session, school, school_class, academic_year)

    resp = await client.get(
        "/students/export/custom?fields=admission_number,current_class,level", headers=auth,
    )
    assert resp.status_code == 200
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8-sig"))))
    header, data_row = rows[0], rows[1]
    assert data_row[header.index("Current Class")] == "3 A"
    assert data_row[header.index("Level")] == "SHS"


# ── Scoping (core/student_scope.py) — regression for a gap where export ──────
# never applied resolve_student_view_scope() at all, unlike list_students.

@pytest.mark.asyncio
async def test_export_csv_scoped_to_class_teacher_own_class(
    client: AsyncClient, auth: dict, db_session: AsyncSession, school: School,
    school_class: Class, academic_year: AcademicYear, redis_permissions: None,
):
    sid_in_class = (await client.post("/students", json=_student("ADM001"), headers=auth)).json()["id"]
    await client.post("/students", json=_student("ADM002"), headers=auth)  # not in the teacher's class
    await client.post("/students/class-assignments", json={
        "student_id": sid_in_class, "class_id": str(school_class.id),
        "academic_year_id": str(academic_year.id),
    }, headers=auth)

    teacher_auth, staff_id = await _login_as_position(client, auth, db_session, school, "CLASS_TEACHER")

    unassigned = await client.get("/students/export/custom?fields=admission_number&fmt=csv", headers=teacher_auth)
    assert unassigned.status_code == 200
    rows = list(csv.reader(io.StringIO(unassigned.content.decode("utf-8-sig"))))
    assert len(rows) == 1  # header only — no students in scope yet

    db_session.add(ClassTeacher(
        school_id=school.id, class_id=school_class.id, staff_member_id=staff_id,
        academic_year_id=academic_year.id, is_active=True,
    ))
    await db_session.flush()

    resp = await client.get("/students/export/custom?fields=admission_number&fmt=csv", headers=teacher_auth)
    assert resp.status_code == 200
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8-sig"))))
    header, data_row = rows[0], rows[1]
    assert len(rows) == 2  # header + the one student in this teacher's class
    assert data_row[header.index("Admission No.")] == "ADM001"


@pytest.mark.asyncio
async def test_custom_export_scoped_to_class_teacher_own_class(
    client: AsyncClient, auth: dict, db_session: AsyncSession, school: School,
    school_class: Class, academic_year: AcademicYear, redis_permissions: None,
):
    (await client.post("/students", json=_student("ADM001"), headers=auth)).json()["id"]
    teacher_auth, _staff_id = await _login_as_position(client, auth, db_session, school, "CLASS_TEACHER")

    resp = await client.get(
        "/students/export/custom?fields=admission_number", headers=teacher_auth,
    )
    assert resp.status_code == 200
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8-sig"))))
    assert len(rows) == 1  # header only — CLASS_TEACHER has no ClassTeacher rows yet, sees nobody


# ── graduated filter parity with the on-screen list (services/student_query.py) ──
# list_students() supports graduated=true/false (12y); the export paths never
# declared or applied it at all, so filtering the on-screen list by
# "Graduated" and exporting silently ignored the filter — the export
# returned every inactive student, not just genuinely graduated ones.

@pytest.mark.asyncio
async def test_export_csv_respects_graduated_filter(
    client: AsyncClient, auth: dict, db_session: AsyncSession,
    school: School, school_admin: User, academic_year: AcademicYear,
):
    grad_id = (await client.post("/students", json=_student("GRAD001"), headers=auth)).json()["id"]
    other_id = (await client.post("/students", json=_student("OTH001", first_name="Ama", last_name="Boateng"), headers=auth)).json()["id"]
    db_session.add(GraduationRecord(
        school_id=school.id, student_id=uuid.UUID(grad_id), academic_year_id=academic_year.id,
        graduation_type=GraduationType.GRADUATED,
        processed_at=datetime(2026, 7, 1), processed_by_id=school_admin.id,
    ))
    await db_session.flush()
    await client.patch(f"/students/{grad_id}", json={"is_active": False}, headers=auth)
    await client.patch(f"/students/{other_id}", json={"is_active": False}, headers=auth)

    resp = await client.get(
        "/students/export/custom?fields=admission_number&fmt=csv&graduated=true&active_only=false", headers=auth,
    )
    assert resp.status_code == 200
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8-sig"))))
    header, data_rows = rows[0], rows[1:]
    admission_numbers = [r[header.index("Admission No.")] for r in data_rows]
    assert admission_numbers == ["GRAD001"], "export must match exactly what graduated=true matches on screen"


@pytest.mark.asyncio
async def test_custom_export_respects_graduated_filter(
    client: AsyncClient, auth: dict, db_session: AsyncSession,
    school: School, school_admin: User, academic_year: AcademicYear,
):
    grad_id = (await client.post("/students", json=_student("GRAD002"), headers=auth)).json()["id"]
    other_id = (await client.post("/students", json=_student("OTH002", first_name="Ama", last_name="Boateng"), headers=auth)).json()["id"]
    db_session.add(GraduationRecord(
        school_id=school.id, student_id=uuid.UUID(grad_id), academic_year_id=academic_year.id,
        graduation_type=GraduationType.GRADUATED,
        processed_at=datetime(2026, 7, 1), processed_by_id=school_admin.id,
    ))
    await db_session.flush()
    await client.patch(f"/students/{grad_id}", json={"is_active": False}, headers=auth)
    await client.patch(f"/students/{other_id}", json={"is_active": False}, headers=auth)

    resp = await client.get(
        "/students/export/custom?fields=admission_number&graduated=true&active_only=false", headers=auth,
    )
    assert resp.status_code == 200
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8-sig"))))
    header, data_rows = rows[0], rows[1:]
    admission_numbers = [r[header.index("Admission No.")] for r in data_rows]
    assert admission_numbers == ["GRAD002"], "export must match exactly what graduated=true matches on screen"


# ── Format parity with staff export (services/export_utils.py, services/pdf.py) ──

@pytest.mark.asyncio
async def test_export_pdf_basic(client: AsyncClient, auth: dict):
    await client.post("/students", json=_student("ADM001"), headers=auth)
    resp = await client.get(
        "/students/export/custom?fields=admission_number,full_name&fmt=pdf", headers=auth,
    )
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content.startswith(b"%PDF")


@pytest.mark.asyncio
async def test_export_excel_basic(client: AsyncClient, auth: dict):
    from openpyxl import load_workbook
    await client.post("/students", json=_student("ADM001"), headers=auth)
    resp = await client.get(
        "/students/export/custom?fields=admission_number,full_name&fmt=excel", headers=auth,
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    assert rows[1][header.index("Admission No.")] == "ADM001"


@pytest.mark.asyncio
async def test_export_rejects_unknown_format(client: AsyncClient, auth: dict):
    resp = await client.get("/students/export/custom?fields=admission_number&fmt=doc", headers=auth)
    assert resp.status_code == 422
