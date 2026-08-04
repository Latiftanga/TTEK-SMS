"""
Student bulk import integration tests.
Run inside Docker: docker compose exec api pytest app/tests/test_student_import.py -v
"""
import io
import pytest
from httpx import AsyncClient

from app.models.school import School
from app.services.student_import_constants import DATA_START, _COLS, SENTINEL_CELL, make_sentinel

_XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_NUM_COLS = len(_COLS)


def _make_xlsx(rows: list[list], school_code: str) -> bytes:
    """Build a minimal .xlsx with the school-specific sentinel and supplied data rows."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Data"
    ws[SENTINEL_CELL] = make_sentinel(school_code)
    for col, label, *_ in _COLS:
        ws[f"{col}3"] = label
    for r_idx, row_vals in enumerate(rows):
        row_num = DATA_START + r_idx
        # Pad/truncate to exactly _NUM_COLS columns
        padded = list(row_vals) + [None] * _NUM_COLS
        for c_idx in range(_NUM_COLS):
            ws.cell(row=row_num, column=c_idx + 1, value=padded[c_idx])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row(admission, first, last, middle=None, dob=None, gender=None,
         nationality=None, religion=None, hometown=None, address=None,
         nhis=None, ghana_card=None, boarding="No", orphan="None", disability=None):
    return [admission, first, middle, last, dob, gender, nationality, religion,
            hometown, address, nhis, ghana_card, boarding, orphan, disability]


# ── Template download ─────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_download_template_returns_xlsx(client: AsyncClient, auth: dict):
    resp = await client.get("/students/import/template", headers=auth)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(_XLSX_CT)
    assert len(resp.content) > 3_000


@pytest.mark.asyncio
async def test_template_contains_school_sentinel(client: AsyncClient, auth: dict, school: School):
    from openpyxl import load_workbook
    resp = await client.get("/students/import/template", headers=auth)
    wb = load_workbook(io.BytesIO(resp.content), data_only=True)
    ws = wb["Student Data"]
    assert ws[SENTINEL_CELL].value == make_sentinel(school.school_code)


@pytest.mark.asyncio
async def test_template_has_all_headers(client: AsyncClient, auth: dict):
    from openpyxl import load_workbook
    resp = await client.get("/students/import/template", headers=auth)
    wb = load_workbook(io.BytesIO(resp.content), data_only=True)
    ws = wb["Student Data"]
    headers = [ws[f"{col}3"].value for col, *_ in _COLS]
    assert all(h is not None for h in headers)
    assert len(headers) == 15  # A–O


# ── Upload processing ─────────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_import_all_valid(client: AsyncClient, auth: dict, school: School):
    rows = [
        _row("ADM001", "Ama",  "Boateng", gender="FEMALE", boarding="No"),
        _row("ADM002", "Kofi", "Asante",  gender="MALE",   boarding="Yes",
             nhis="GH-99", orphan="Half Orphan"),
    ]
    resp = await client.post(
        "/students/import",
        files={"file": ("students.xlsx", _make_xlsx(rows, school.school_code), _XLSX_CT)},
        headers=auth,
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["created"] == 2
    assert data["failed"] == 0
    assert data["errors"] == []
    assert "batch_id" in data


@pytest.mark.asyncio
async def test_import_partial_success(client: AsyncClient, auth: dict, school: School):
    """Two valid rows + one missing last_name → 2 created, 1 failed."""
    rows = [
        _row("ADM001", "Ama",  "Boateng"),
        _row("ADM002", "Kofi", "Asante"),
        _row("ADM003", "Yaw",  None),       # missing last_name
    ]
    resp = await client.post(
        "/students/import",
        files={"file": ("students.xlsx", _make_xlsx(rows, school.school_code), _XLSX_CT)},
        headers=auth,
    )
    data = resp.json()
    assert data["created"] == 2
    assert data["failed"] == 1
    assert data["errors"][0]["row"] == DATA_START + 2


@pytest.mark.asyncio
async def test_import_duplicate_admission_number(client: AsyncClient, auth: dict, school: School):
    """Duplicate admission number in same file → first created, second failed."""
    rows = [
        _row("ADM001", "Ama",  "Boateng"),
        _row("ADM001", "Akua", "Mensah"),
    ]
    resp = await client.post(
        "/students/import",
        files={"file": ("students.xlsx", _make_xlsx(rows, school.school_code), _XLSX_CT)},
        headers=auth,
    )
    data = resp.json()
    assert data["created"] == 1
    assert data["failed"] == 1
    assert "already exists" in data["errors"][0]["error"]


@pytest.mark.asyncio
async def test_import_blank_admission_number_rejected_with_clear_message(
    client: AsyncClient, auth: dict, school: School,
):
    """The template marks admission_number required (sheet header + Instructions
    tab), but StudentCreate itself allows it to be omitted (the single-create
    endpoint auto-generates one) — a blank cell must be reported as a clear
    validation failure, not fall through to the DB's NOT NULL constraint and
    come back mislabeled as a duplicate."""
    rows = [
        _row("ADM001", "Ama", "Boateng"),
        _row(None, "Yaw", "Mensah"),  # blank admission number, otherwise valid
    ]
    resp = await client.post(
        "/students/import",
        files={"file": ("students.xlsx", _make_xlsx(rows, school.school_code), _XLSX_CT)},
        headers=auth,
    )
    data = resp.json()
    assert data["created"] == 1
    assert data["failed"] == 1
    assert "required" in data["errors"][0]["error"].lower()
    assert "already exists" not in data["errors"][0]["error"]


@pytest.mark.asyncio
async def test_import_rejects_existing_admission_number(client: AsyncClient, auth: dict, school: School):
    """Admission number already in DB is rejected."""
    await client.post("/students", json={
        "admission_number": "PRE001", "first_name": "Existing", "last_name": "Student",
    }, headers=auth)
    rows = [_row("PRE001", "New", "Person")]
    resp = await client.post(
        "/students/import",
        files={"file": ("students.xlsx", _make_xlsx(rows, school.school_code), _XLSX_CT)},
        headers=auth,
    )
    data = resp.json()
    assert data["created"] == 0
    assert data["failed"] == 1


@pytest.mark.asyncio
async def test_import_cross_school_sentinel_rejected(client: AsyncClient, auth: dict):
    """Template from a different school returns 422 with a helpful message."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Data"
    ws[SENTINEL_CELL] = make_sentinel("OTHER_SCHOOL")
    buf = io.BytesIO()
    wb.save(buf)
    resp = await client.post(
        "/students/import",
        files={"file": ("other.xlsx", buf.getvalue(), _XLSX_CT)},
        headers=auth,
    )
    assert resp.status_code == 422
    assert "OTHER_SCHOOL" in resp.json()["detail"]


@pytest.mark.asyncio
async def test_import_wrong_template_rejected(client: AsyncClient, auth: dict):
    """Generic .xlsx with no sentinel → 422."""
    from openpyxl import Workbook
    wb = Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    resp = await client.post(
        "/students/import",
        files={"file": ("generic.xlsx", buf.getvalue(), _XLSX_CT)},
        headers=auth,
    )
    assert resp.status_code == 422


@pytest.mark.asyncio
async def test_import_non_xlsx_rejected(client: AsyncClient, auth: dict):
    resp = await client.post(
        "/students/import",
        files={"file": ("students.csv", b"admission_number,first_name\nADM001,Ama", "text/csv")},
        headers=auth,
    )
    assert resp.status_code == 422


@pytest.mark.asyncio
async def test_import_new_fields_saved(client: AsyncClient, auth: dict, school: School):
    """NHIS, Ghana Card, boarding, orphan status, disability all persisted."""
    rows = [_row(
        "ADM999", "Kwame", "Acheampong",
        nhis="GH-00112233", ghana_card="GHA-999888777",
        boarding="Yes", orphan="Full Orphan", disability="Hearing impairment",
    )]
    resp = await client.post(
        "/students/import",
        files={"file": ("students.xlsx", _make_xlsx(rows, school.school_code), _XLSX_CT)},
        headers=auth,
    )
    assert resp.status_code == 200
    assert resp.json()["created"] == 1

    # Fetch the student and check the new fields
    list_resp = await client.get("/students?search=ADM999", headers=auth)
    students = list_resp.json()
    assert len(students) == 1
    detail_resp = await client.get(f"/students/{students[0]['id']}", headers=auth)
    s = detail_resp.json()
    assert s["nhis_number"] == "GH-00112233"
    assert s["ghana_card_number"] == "GHA-999888777"
    assert s["is_boarding"] is True
    assert s["orphan_status"] == "FULL_ORPHAN"
    assert s["disability"] == "Hearing impairment"
