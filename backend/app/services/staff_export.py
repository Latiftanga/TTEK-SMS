"""Staff register export — Excel (openpyxl) and PDF (WeasyPrint).

Used by GES/CAGD staff data submissions.  The rank column is omitted for
PRIVATE schools because they do not use GES promotions.

Current rank is derived at query time from the latest promotion's to_rank.title.
"""
from __future__ import annotations
import io
import uuid
from datetime import date, datetime, timezone

from sqlalchemy import select
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.school import School, SchoolOwnership
from app.models.staff import StaffMember
from app.models.staff_history import StaffPromotion
from app.services.staff_query import staff_search_condition


async def _load_staff(
    school_id: uuid.UUID,
    db: AsyncSession,
    category_id: uuid.UUID | None,
    active_only: bool,
    search: str | None = None,
    gender: str | None = None,
) -> tuple[School, list[StaffMember]]:
    school = await db.get(School, school_id)
    q = (
        select(StaffMember)
        .where(StaffMember.school_id == school_id)
        .options(
            selectinload(StaffMember.category),
            selectinload(StaffMember.promotions).selectinload(StaffPromotion.to_rank),
        )
        .order_by(StaffMember.last_name, StaffMember.first_name)
    )
    if active_only:
        q = q.where(StaffMember.is_active == True)
    if category_id:
        q = q.where(StaffMember.category_id == category_id)
    if gender:
        q = q.where(StaffMember.gender == gender)
    if search:
        q = q.where(staff_search_condition(search))
    members = list(await db.scalars(q))
    return school, members


def _current_rank(member: StaffMember) -> str:
    if not member.promotions:
        return "—"
    latest = sorted(member.promotions, key=lambda p: p.effective_date, reverse=True)[0]
    return latest.to_rank.title if latest.to_rank else "—"


async def export_excel(
    school_id: uuid.UUID,
    db: AsyncSession,
    category_id: uuid.UUID | None = None,
    active_only: bool = True,
    search: str | None = None,
    gender: str | None = None,
) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is not installed. Add it to requirements.txt.") from exc

    school, members = await _load_staff(school_id, db, category_id, active_only, search=search, gender=gender)
    show_rank = school and school.ownership == SchoolOwnership.PUBLIC

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Register"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    center = Alignment(horizontal="center")

    headers = ["No.", "Full Name", "Staff No.", "Category"]
    if show_rank:
        headers.append("Current Rank")
    headers += ["SSNIT No.", "Ghana Card", "Employment Type", "Date Joined"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, m in enumerate(members, 2):
        cols = [
            row_idx - 1,
            f"{m.last_name}, {m.first_name}{(' ' + m.middle_name) if m.middle_name else ''}",
            m.staff_number,
            m.category.name if m.category else "—",
        ]
        if show_rank:
            cols.append(_current_rank(m))
        cols += [
            m.ssnit_number or "—",
            m.national_id or "—",
            m.employment_type.value if m.employment_type else "—",
            m.joined_date.isoformat() if m.joined_date else "—",
        ]
        for col_idx, val in enumerate(cols, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def export_pdf(
    school_id: uuid.UUID,
    db: AsyncSession,
    category_id: uuid.UUID | None = None,
    active_only: bool = True,
    search: str | None = None,
    gender: str | None = None,
) -> bytes:
    from jinja2 import Environment, FileSystemLoader
    from weasyprint import HTML
    import os

    school, members = await _load_staff(school_id, db, category_id, active_only, search=search, gender=gender)
    show_rank = school and school.ownership == SchoolOwnership.PUBLIC

    template_dir = os.path.join(os.path.dirname(__file__), "..", "templates")
    env = Environment(loader=FileSystemLoader(os.path.abspath(template_dir)), autoescape=True)
    tmpl = env.get_template("staff_register.html")

    rows = []
    for i, m in enumerate(members, 1):
        row: dict = {
            "no": i,
            "name": f"{m.last_name}, {m.first_name}{(' ' + m.middle_name) if m.middle_name else ''}",
            "staff_number": m.staff_number,
            "category": m.category.name if m.category else "—",
            "ssnit": m.ssnit_number or "—",
            "ghana_card": m.national_id or "—",
            "employment_type": m.employment_type.value if m.employment_type else "—",
            "joined_date": m.joined_date.isoformat() if m.joined_date else "—",
        }
        if show_rank:
            row["current_rank"] = _current_rank(m)
        rows.append(row)

    html_str = tmpl.render(
        school=school,
        rows=rows,
        show_rank=show_rank,
        generated_at=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        total=len(rows),
    )
    return HTML(string=html_str).write_pdf()
