"""Custom-field student export — CSV or Excel.

Caller selects which columns to include from STUDENT_FIELDS.
All standard filter params (search, class_id, gender, level, etc.) are respected.
"""
from __future__ import annotations
import csv
import io
import uuid
from typing import Callable

from sqlalchemy import or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.academic import Class
from app.models.students import (
    Guardian, Student, StudentClassAssignment, StudentGuardian, TermEnrollment,
)
from app.services.student import _class_display, _display_name, _get_class_map

# Keys of fields that require the class or guardian lookup queries.
_CLASS_FIELDS    = {"current_class", "level"}
_GUARDIAN_FIELDS = {"guardian_name", "guardian_phone"}

# (header_label, extractor(student, class_map, guardian_map) -> str)
type _Extractor = Callable[[Student, dict, dict], str]

STUDENT_FIELDS: dict[str, tuple[str, _Extractor]] = {
    "admission_number":  ("Admission No.",   lambda s, c, g: s.admission_number),
    "full_name":         ("Full Name",        lambda s, c, g: _display_name(s.first_name, s.middle_name, s.last_name)),
    "first_name":        ("First Name",       lambda s, c, g: s.first_name),
    "last_name":         ("Last Name",        lambda s, c, g: s.last_name),
    "middle_name":       ("Middle Name",      lambda s, c, g: s.middle_name or ""),
    "gender":            ("Gender",           lambda s, c, g: s.gender.value if s.gender else ""),
    "date_of_birth":     ("Date of Birth",    lambda s, c, g: str(s.date_of_birth) if s.date_of_birth else ""),
    "nationality":       ("Nationality",      lambda s, c, g: s.nationality or ""),
    "religion":          ("Religion",         lambda s, c, g: s.religion or ""),
    "hometown":          ("Hometown",         lambda s, c, g: s.hometown or ""),
    "current_class":     ("Current Class",    lambda s, c, g: c.get(s.id, ("", ""))[0]),
    "level":             ("Level",            lambda s, c, g: c.get(s.id, ("", ""))[1]),
    "boarding":          ("Boarding",         lambda s, c, g: "Boarding" if s.is_boarding else "Day"),
    "nhis_number":       ("NHIS No.",         lambda s, c, g: s.nhis_number or ""),
    "ghana_card_number": ("Ghana Card No.",   lambda s, c, g: s.ghana_card_number or ""),
    "orphan_status":     ("Orphan Status",    lambda s, c, g: s.orphan_status.value if s.orphan_status else ""),
    "disability":        ("Disability",       lambda s, c, g: s.disability or ""),
    "guardian_name":     ("Guardian Name",    lambda s, c, g: g.get(s.id, ("", ""))[0]),
    "guardian_phone":    ("Guardian Phone",   lambda s, c, g: g.get(s.id, ("", ""))[1]),
    "status":            ("Status",           lambda s, c, g: "Active" if s.is_active else "Inactive"),
}


async def export_students_custom(
    school_id: uuid.UUID,
    db: AsyncSession,
    *,
    fields: list[str],
    fmt: str = "csv",
    active_only: bool = True,
    class_id: uuid.UUID | None = None,
    term_id: uuid.UUID | None = None,
    gender: str | None = None,
    search: str | None = None,
    level: str | None = None,
    year_group: int | None = None,
) -> bytes:
    # Resolve requested field keys against registry (preserve order, skip unknown)
    resolved = [(k, STUDENT_FIELDS[k]) for k in fields if k in STUDENT_FIELDS]
    if not resolved:
        resolved = list(STUDENT_FIELDS.items())

    # ── Query ─────────────────────────────────────────────────────────────────
    q = select(Student).where(Student.school_id == school_id)
    if active_only:
        q = q.where(Student.is_active == True)  # noqa: E712
    if gender:
        q = q.where(Student.gender == gender)
    if class_id or level or year_group:
        q = q.join(StudentClassAssignment, StudentClassAssignment.student_id == Student.id).where(
            StudentClassAssignment.is_active == True,  # noqa: E712
        )
        if class_id:
            q = q.where(StudentClassAssignment.class_id == class_id)
        if level or year_group:
            q = q.join(Class, Class.id == StudentClassAssignment.class_id)
            if level:
                q = q.where(Class.level == level)
            if year_group:
                q = q.where(Class.year_group == year_group)
    if term_id:
        q = q.join(TermEnrollment, TermEnrollment.student_id == Student.id).where(
            TermEnrollment.is_active == True,  # noqa: E712
            TermEnrollment.academic_term_id == term_id,
        )
    if search:
        pat = f"%{search}%"
        q = q.where(or_(
            Student.first_name.ilike(pat),
            Student.last_name.ilike(pat),
            Student.admission_number.ilike(pat),
        ))
    students = list(await db.scalars(q.distinct().order_by(Student.last_name, Student.first_name)))
    sids = [s.id for s in students]

    # ── Ancillary lookups (only when needed) ─────────────────────────────────
    requested_keys = {k for k, _ in resolved}
    # id → (display, level) — resolves to the most-recent active assignment (see
    # _active_class_assignment_subquery's docstring for why a promoted student
    # can have 2+ active rows and why "first row seen" would be non-deterministic).
    class_map: dict[uuid.UUID, tuple[str, str]] = {}
    if sids and requested_keys & _CLASS_FIELDS:
        class_info_map = await _get_class_map(sids, db)
        class_map = {
            sid: (_class_display(level, year_group, programme, stream), level)
            for sid, (level, year_group, programme, stream, _cls_id) in class_info_map.items()
        }

    guardian_map: dict[uuid.UUID, tuple[str, str]] = {}
    if sids and requested_keys & _GUARDIAN_FIELDS:
        g_rows = await db.execute(
            select(StudentGuardian.student_id, Guardian.first_name, Guardian.last_name, Guardian.phone)
            .join(Guardian, Guardian.id == StudentGuardian.guardian_id)
            .where(StudentGuardian.student_id.in_(sids), StudentGuardian.is_primary == True)  # noqa: E712
        )
        for gr in g_rows:
            guardian_map[gr.student_id] = (f"{gr.first_name} {gr.last_name}", gr.phone)

    # ── Build rows ────────────────────────────────────────────────────────────
    headers = [label for _, (label, _ext) in resolved]
    data_rows = [
        [extractor(s, class_map, guardian_map) for _, (_lbl, extractor) in resolved]
        for s in students
    ]

    return _to_bytes(headers, data_rows, fmt)


def _to_bytes(headers: list[str], rows: list[list[str]], fmt: str) -> bytes:
    if fmt == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise RuntimeError("openpyxl is not installed.") from exc
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        hfill = PatternFill("solid", fgColor="D9E1F2")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col) + 4, 50
            )
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")
