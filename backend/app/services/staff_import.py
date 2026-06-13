"""
Staff bulk import processing.

process_import() reads a .xlsx file produced by build_template(), validates
every data row with Pydantic, and inserts valid rows using per-row savepoints
so a single failure never aborts the whole batch (best-effort strategy).

Returned ImportBatchResult contains:
  - batch_id: the ImportBatch record id for audit purposes
  - total_rows / created / failed: summary counts
  - errors: list of per-row failures with row number, staff_number, and reason
"""
from __future__ import annotations
import io
import uuid
from datetime import date, datetime, timezone

from fastapi import HTTPException
from pydantic import ValidationError
from sqlalchemy import or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.auth import StaffPosition
from app.models.documents import ImportBatch, ImportRow, ImportStatus
from app.models.staff import StaffMember
from app.schemas.documents import ImportBatchResult, ImportRowResult
from app.schemas.staff import StaffMemberCreate
from app.services.staff_import_constants import _COLS, DATA_START, make_sentinel


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _parse_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def _str(cell) -> str | None:
    v = cell.value
    return str(v).strip() or None if v is not None else None


async def _load_position_map(school_id: uuid.UUID, db: AsyncSession) -> dict[str, uuid.UUID]:
    rows = await db.scalars(
        select(StaffPosition).where(
            or_(StaffPosition.school_id == school_id, StaffPosition.school_id.is_(None))
        )
    )
    return {p.name.lower(): p.id for p in rows}


async def process_import(
    file_bytes: bytes,
    school_id: uuid.UUID,
    school_code: str,
    user_id: uuid.UUID,
    db: AsyncSession,
) -> ImportBatchResult:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="Cannot read file — ensure it is a valid .xlsx.")

    ws = wb["Staff Data"] if "Staff Data" in wb.sheetnames else wb.active
    sentinel = ws["N1"].value
    expected = make_sentinel(school_code)
    if sentinel != expected:
        if sentinel and str(sentinel).startswith("TTEK_STAFF_IMPORT_"):
            embedded = str(sentinel).removeprefix("TTEK_STAFF_IMPORT_")
            detail = (
                f"This template was generated for school '{embedded}' — "
                f"you are logged in as '{school_code.upper()}'. "
                "Please download a fresh template for your school."
            )
        else:
            detail = "Unrecognised template. Download the official template from GET /staff/import/template."
        raise HTTPException(status_code=422, detail=detail)

    pos_map = await _load_position_map(school_id, db)

    batch = ImportBatch(
        school_id=school_id, import_type="staff", status=ImportStatus.PROCESSING,
        total_rows=0, processed_rows=0, error_count=0,
        initiated_by_id=user_id, started_at=_utcnow(), created_at=_utcnow(),
    )
    db.add(batch)
    await db.flush()

    results: list[ImportRowResult] = []
    created = failed = 0

    for row_num in range(DATA_START, (ws.max_row or DATA_START) + 1):
        cells = {col: ws[f"{col}{row_num}"] for col, *_ in _COLS}
        sn = _str(cells["A"])
        fn = _str(cells["B"])
        ln = _str(cells["D"])
        if not any([sn, fn, ln]):
            continue

        raw: dict = {}
        for col, _, field, _, _ in _COLS:
            if field in ("date_of_birth", "joined_date"):
                raw[field] = _parse_date(cells[col].value)
            else:
                raw[field] = _str(cells[col])

        position_name = raw.pop("position_name", None)
        raw["position_id"] = pos_map.get(position_name.lower()) if position_name else None

        try:
            req = StaffMemberCreate(**raw)
        except ValidationError as exc:
            msg = "; ".join(f"{e['loc'][-1]}: {e['msg']}" for e in exc.errors())
            _log_row(db, batch.id, school_id, row_num, raw, "error", msg)
            results.append(ImportRowResult(row=row_num, ref=sn, status="failed", error=msg))
            failed += 1
            continue

        member = StaffMember(
            school_id=school_id,
            staff_number=req.staff_number,
            first_name=req.first_name,
            middle_name=req.middle_name,
            last_name=req.last_name,
            date_of_birth=req.date_of_birth,
            gender=req.gender,
            national_id=req.national_id,
            phone=req.phone,
            email=req.email.lower().strip() if req.email else None,
            position_id=req.position_id,
            department=req.department,
            is_active=True,
            joined_date=req.joined_date,
        )
        try:
            async with db.begin_nested():
                db.add(member)
                await db.flush()
            _log_row(db, batch.id, school_id, row_num, raw, "success", None, member.id)
            results.append(ImportRowResult(row=row_num, ref=req.staff_number, status="created", error=None))
            created += 1
        except IntegrityError:
            try:
                db.expunge(member)
            except Exception:
                pass
            msg = f"Staff number '{req.staff_number}' already exists at this school."
            _log_row(db, batch.id, school_id, row_num, raw, "error", msg)
            results.append(ImportRowResult(row=row_num, ref=req.staff_number, status="failed", error=msg))
            failed += 1

    batch.total_rows = created + failed
    batch.processed_rows = created + failed
    batch.error_count = failed
    batch.status = (
        ImportStatus.COMPLETED if failed == 0
        else ImportStatus.PARTIAL if created > 0
        else ImportStatus.FAILED
    )
    batch.completed_at = _utcnow()
    await db.flush()

    return ImportBatchResult(
        batch_id=batch.id,
        total_rows=created + failed,
        created=created,
        failed=failed,
        errors=[r for r in results if r.status == "failed"],
    )


def _log_row(
    db: AsyncSession,
    batch_id: uuid.UUID,
    school_id: uuid.UUID,
    row_num: int,
    raw: dict,
    status: str,
    error_message: str | None,
    entity_id: uuid.UUID | None = None,
) -> None:
    db.add(ImportRow(
        school_id=school_id,
        batch_id=batch_id,
        row_number=row_num,
        raw_data={k: str(v) if v is not None else None for k, v in raw.items()},
        status=status,
        error_message=error_message,
        entity_id=entity_id,
    ))
