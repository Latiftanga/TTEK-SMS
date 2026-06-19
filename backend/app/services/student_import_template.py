"""
Branded Excel template generation for student bulk import.

Called by:  GET /students/import/template
Returns:    .xlsx bytes ready to stream to the browser

Structure:
  Row 1: Merged title with school name and brand_color
  Row 2: Subtitle / instructions summary
  Row 3: Column headers (locked)
  Row 4: Sample/example row (greyed, italic)
  Rows 5-204: Pre-formatted data-entry area with dropdowns
  Sentinel cell Q1 = "TTEK_STUDENT_IMPORT_{school_code}"
  "Instructions" sheet with field guide
"""
from __future__ import annotations
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation

from app.services.student_import_constants import (
    _COLS, _DATA_START, _NUM_ROWS, SENTINEL_CELL, make_sentinel,
)


def _argb(hex_color: str) -> str:
    return "FF" + hex_color.lstrip("#").upper()


def build_template(school_name: str, school_code: str, brand_color: str) -> bytes:
    """Return branded .xlsx bytes for the student import template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Data"

    brand_fill  = PatternFill("solid", fgColor=_argb(brand_color))
    grey_fill   = PatternFill("solid", fgColor="FFD9D9D9")
    sample_fill = PatternFill("solid", fgColor="FFF0F0F0")
    stripe_fill = PatternFill("solid", fgColor="FFE8EEF8")
    white_fill  = PatternFill("solid", fgColor="FFFFFFFF")
    hdr_font    = Font(bold=True, color="FFFFFFFF", name="Calibri", size=10)
    sample_font = Font(italic=True, color="FF888888", name="Calibri", size=10)
    body_font   = Font(name="Calibri", size=10)
    last_col    = _COLS[-1][0]

    # Row 1 — title
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = f"{school_name}  —  Student Import Template"
    c.font  = Font(bold=True, color="FFFFFFFF", name="Calibri", size=14)
    c.fill  = brand_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    ws[SENTINEL_CELL].value = make_sentinel(school_code)

    # Row 2 — subtitle
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = (
        "Enter student details from row 5 onward.  Required fields are marked *.  "
        "You may delete or overwrite the example row (row 4) before uploading."
    )
    c.font  = Font(italic=True, color="FF555555", name="Calibri", size=9)
    c.fill  = grey_fill
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    # Row 3 — column headers
    for col, label, _, _, width in _COLS:
        c = ws[f"{col}3"]
        c.value = label
        c.font  = hdr_font
        c.fill  = brand_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col].width = width
    ws.row_dimensions[3].height = 28

    # Row 4 — sample row
    sample_vals = [
        "ADM2024001", "Ama", "Akua", "Boateng",
        "2006-03-15", "FEMALE", "Ghanaian", "Christian", "Kumasi",
        "House No. 5, Adum, Kumasi",
        "GH-12345678", "GHA-000011223",
        "No", "None", "",
    ]
    for (col, *_), val in zip(_COLS, sample_vals):
        c = ws[f"{col}4"]
        c.value = val
        c.font  = sample_font
        c.fill  = sample_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 18

    # Rows 5+ — data entry area
    for row in range(_DATA_START, _DATA_START + _NUM_ROWS):
        fill = white_fill if row % 2 == 0 else stripe_fill
        for col, *_ in _COLS:
            c = ws[f"{col}{row}"]
            c.fill = fill
            c.font = body_font
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.protection = Protection(locked=False)
        ws.row_dimensions[row].height = 17

    data_range_end = _DATA_START + _NUM_ROWS - 1

    # Gender dropdown (col F)
    dv_gender = DataValidation(type="list", formula1='"MALE,FEMALE"', allow_blank=True)
    dv_gender.sqref = f"F{_DATA_START}:F{data_range_end}"
    ws.add_data_validation(dv_gender)

    # Boarding dropdown (col M)
    dv_boarding = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_boarding.sqref = f"M{_DATA_START}:M{data_range_end}"
    ws.add_data_validation(dv_boarding)

    # Orphan status dropdown (col N)
    dv_orphan = DataValidation(
        type="list",
        formula1='"None,Half Orphan,Full Orphan"',
        allow_blank=True,
    )
    dv_orphan.sqref = f"N{_DATA_START}:N{data_range_end}"
    ws.add_data_validation(dv_orphan)

    ws.freeze_panes = f"A{_DATA_START}"
    ws.protection.sheet = True
    ws.protection.password = ""

    _build_instructions_sheet(wb, brand_fill, hdr_font)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_instructions_sheet(wb: Workbook, brand_fill: PatternFill, hdr_font: Font) -> None:
    ws = wb.create_sheet("Instructions")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 68
    rows = [
        ("TTEK Student Import — Field Guide", None, True),
        (None, None, False),
        ("Field", "Notes / Valid Values", True),
        ("Admission Number*",       "Unique ID for this student at this school. Required.",         False),
        ("First Name*",             "Required.",                                                     False),
        ("Middle Name",             "Optional.",                                                     False),
        ("Last Name*",              "Required.",                                                     False),
        ("Date of Birth",           "Format: YYYY-MM-DD  (e.g. 2006-03-15)",                        False),
        ("Gender",                  "Select MALE or FEMALE from the dropdown.",                      False),
        ("Nationality",             "e.g. Ghanaian, Nigerian.",                                      False),
        ("Religion",                "e.g. Christian, Muslim, Traditional.",                         False),
        ("Hometown",                "Town or city of origin.",                                       False),
        ("Residential Address",     "Full address while at school.",                                 False),
        ("NHIS Number",             "National Health Insurance Scheme number (optional).",           False),
        ("Ghana Card Number",       "Ghana Card / National ID number (optional).",                   False),
        ("Boarding (Yes/No)",       "Select Yes if the student lives in a school house; else No.",   False),
        ("Orphan Status",           "Select None, Half Orphan, or Full Orphan from the dropdown.",  False),
        ("Disability/Special Needs","Brief description of any disability or special need (optional).", False),
        (None, None, False),
        ("Tips", "", True),
        ("", "• Do not edit or delete header rows 1–3.",                                             False),
        ("", "• You may delete or overwrite the example row (row 4).",                              False),
        ("", "• Leave a cell blank if the information is not yet known.",                            False),
        ("", "• Save the file as .xlsx before uploading.",                                           False),
        ("", "• Duplicate admission numbers within the same school are rejected.",                   False),
    ]
    for i, (a, b, is_hdr) in enumerate(rows, 1):
        ca = ws.cell(row=i, column=1, value=a)
        cb = ws.cell(row=i, column=2, value=b)
        if is_hdr and a:
            ca.font = cb.font = hdr_font
            ca.fill = cb.fill = brand_fill
            if i == 1:
                ws.merge_cells("A1:B1")
                ws.row_dimensions[1].height = 24
