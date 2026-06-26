"""Custom-field staff export — CSV or Excel.

Caller selects which columns to include from STAFF_FIELDS.
All standard filter params (search, gender, category_id, active_only) are respected.
"""
from __future__ import annotations
import csv
import io
import uuid
from typing import Callable

from sqlalchemy import or_, select
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.school import School, SchoolOwnership
from app.models.staff import StaffMember, StaffPromotion


def _full_name(m: StaffMember) -> str:
    parts = [m.last_name, m.first_name]
    if m.middle_name:
        parts.append(m.middle_name)
    return ", ".join(parts[:1]) + " " + " ".join(parts[1:])


def _current_rank(m: StaffMember) -> str:
    if not m.promotions:
        return "—"
    latest = sorted(m.promotions, key=lambda p: p.effective_date, reverse=True)[0]
    return latest.to_rank.title if latest.to_rank else "—"


type _Extractor = Callable[[StaffMember], str]

STAFF_FIELDS: dict[str, tuple[str, _Extractor]] = {
    "staff_number":    ("Staff No.",        lambda m: m.staff_number),
    "full_name":       ("Full Name",         lambda m: _full_name(m)),
    "first_name":      ("First Name",        lambda m: m.first_name),
    "last_name":       ("Last Name",         lambda m: m.last_name),
    "gender":          ("Gender",            lambda m: m.gender.value if m.gender else ""),
    "date_of_birth":   ("Date of Birth",     lambda m: str(m.date_of_birth) if m.date_of_birth else ""),
    "category":        ("Job Class",         lambda m: m.category.name if m.category else ""),
    "positions":       ("Position(s)",       lambda m: "; ".join(p.name for p in m.positions)),
    "employment_type": ("Employment Type",   lambda m: m.employment_type.value if m.employment_type else ""),
    "joined_date":     ("Date Joined",       lambda m: m.joined_date.isoformat() if m.joined_date else ""),
    "current_rank":    ("Current Rank",      lambda m: _current_rank(m)),
    "ssnit_number":    ("SSNIT No.",         lambda m: m.ssnit_number or ""),
    "national_id":     ("Ghana Card No.",    lambda m: m.national_id or ""),
    "phone":           ("Phone",             lambda m: m.phone or ""),
    "email":           ("Email",             lambda m: m.email or ""),
    "status":          ("Status",            lambda m: "Active" if m.is_active else "Inactive"),
}


async def export_staff_custom(
    school_id: uuid.UUID,
    db: AsyncSession,
    *,
    fields: list[str],
    fmt: str = "csv",
    active_only: bool = True,
    category_id: uuid.UUID | None = None,
    search: str | None = None,
    gender: str | None = None,
) -> bytes:
    resolved = [(k, STAFF_FIELDS[k]) for k in fields if k in STAFF_FIELDS]
    if not resolved:
        resolved = list(STAFF_FIELDS.items())

    q = (
        select(StaffMember)
        .where(StaffMember.school_id == school_id)
        .options(
            selectinload(StaffMember.category),
            selectinload(StaffMember.positions),
            selectinload(StaffMember.promotions).selectinload(StaffPromotion.to_rank),
        )
        .order_by(StaffMember.last_name, StaffMember.first_name)
    )
    if active_only:
        q = q.where(StaffMember.is_active == True)  # noqa: E712
    if category_id:
        q = q.where(StaffMember.category_id == category_id)
    if gender:
        q = q.where(StaffMember.gender == gender)
    if search:
        pat = f"%{search}%"
        q = q.where(or_(
            StaffMember.first_name.ilike(pat),
            StaffMember.last_name.ilike(pat),
            StaffMember.staff_number.ilike(pat),
        ))
    members = list(await db.scalars(q))

    headers = [label for _, (label, _) in resolved]
    data_rows = [
        [extractor(m) for _, (_lbl, extractor) in resolved]
        for m in members
    ]

    return _to_bytes(headers, data_rows, fmt)


def _to_bytes(headers: list[str], rows: list[list[str]], fmt: str) -> bytes:
    if fmt == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise RuntimeError("openpyxl is not installed.") from exc
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Staff"
        hfill = PatternFill("solid", fgColor="D9E1F2")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col) + 4, 50
            )
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")
