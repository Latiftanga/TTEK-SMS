"""Shared CSV/Excel byte-writer for custom exports.

Used by staff_custom_export.py and student_custom_export.py so the two
modules' "resolve fields → build rows → serialize" custom-export flow
shares one writer instead of two byte-for-byte duplicated openpyxl/csv
implementations. PDF output is handled separately by
services/pdf.py::render_export_table() — it needs a School + title, which
this module has no reason to know about.
"""
from __future__ import annotations
import csv
import io


def rows_to_bytes(headers: list[str], rows: list[list[str]], fmt: str, *, sheet_title: str = "Export") -> bytes:
    if fmt == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise RuntimeError("openpyxl is not installed.") from exc
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title
        hfill = PatternFill("solid", fgColor="D9E1F2")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center")
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col) + 4, 50
            )
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")
