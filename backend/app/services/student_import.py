"""
Student bulk import processing.

process_import() reads a .xlsx file produced by build_template(), validates
each data row with Pydantic, and inserts valid rows using per-row savepoints.
Best-effort: failed rows are reported, valid rows are committed.
"""
from __future__ import annotations
import io
import uuid
from datetime import date, datetime, timezone

from fastapi import HTTPException
from pydantic import ValidationError
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.documents import ImportBatch, ImportRow, ImportStatus
from app.models.students import Student
from app.schemas.documents import ImportBatchResult, ImportRowResult
from app.schemas.students import StudentCreate, ORPHAN_STATUSES
from app.services.student_import_constants import (
    _COLS, DATA_START, SENTINEL_CELL, make_sentinel,
)


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _parse_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def _str(cell) -> str | None:
    v = cell.value
    return str(v).strip() or None if v is not None else None


def _parse_bool(cell) -> bool:
    """Parse Yes/No/True/False/1/0 to bool; default False."""
    v = _str(cell)
    if v is None:
        return False
    return v.lower() in ("yes", "true", "1", "y")


def _parse_orphan(cell) -> str:
    """Map human-readable dropdown value to ORPHAN_STATUS enum string."""
    v = _str(cell)
    if v is None:
        return "NONE"
    mapping = {
        "half orphan": "HALF_ORPHAN",
        "full orphan": "FULL_ORPHAN",
        "half_orphan": "HALF_ORPHAN",
        "full_orphan": "FULL_ORPHAN",
    }
    return mapping.get(v.lower(), "NONE")


async def process_import(
    file_bytes: bytes,
    school_id: uuid.UUID,
    school_code: str,
    user_id: uuid.UUID,
    db: AsyncSession,
) -> ImportBatchResult:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="Cannot read file — ensure it is a valid .xlsx.")

    ws = wb["Student Data"] if "Student Data" in wb.sheetnames else wb.active
    sentinel = ws[SENTINEL_CELL].value
    expected = make_sentinel(school_code)
    if sentinel != expected:
        if sentinel and str(sentinel).startswith("TTEK_STUDENT_IMPORT_"):
            embedded = str(sentinel).removeprefix("TTEK_STUDENT_IMPORT_")
            detail = (
                f"This template was generated for school '{embedded}' — "
                f"you are logged in as '{school_code.upper()}'. "
                "Please download a fresh template for your school."
            )
        else:
            detail = "Unrecognised template. Download the official template from GET /students/import/template."
        raise HTTPException(status_code=422, detail=detail)

    batch = ImportBatch(
        school_id=school_id, import_type="student", status=ImportStatus.PROCESSING,
        total_rows=0, processed_rows=0, error_count=0,
        initiated_by_id=user_id, started_at=_utcnow(), created_at=_utcnow(),
    )
    db.add(batch)
    await db.flush()

    results: list[ImportRowResult] = []
    created = failed = 0

    for row_num in range(DATA_START, (ws.max_row or DATA_START) + 1):
        cells = {col: ws[f"{col}{row_num}"] for col, *_ in _COLS}
        an = _str(cells["A"])
        fn = _str(cells["B"])
        ln = _str(cells["D"])
        if not any([an, fn, ln]):
            continue

        raw: dict = {}
        for col, _, field, _, _ in _COLS:
            if field == "date_of_birth":
                raw[field] = _parse_date(cells[col].value)
            elif field == "is_boarding":
                raw[field] = _parse_bool(cells[col])
            elif field == "orphan_status":
                raw[field] = _parse_orphan(cells[col])
            else:
                raw[field] = _str(cells[col])

        try:
            req = StudentCreate(**raw)
        except ValidationError as exc:
            msg = "; ".join(f"{e['loc'][-1]}: {e['msg']}" for e in exc.errors())
            _log_row(db, batch.id, school_id, row_num, raw, "error", msg)
            results.append(ImportRowResult(row=row_num, ref=an, status="failed", error=msg))
            failed += 1
            continue

        student = Student(
            school_id=school_id,
            admission_number=req.admission_number,
            first_name=req.first_name,
            middle_name=req.middle_name,
            last_name=req.last_name,
            date_of_birth=req.date_of_birth,
            gender=req.gender,
            nationality=req.nationality,
            religion=req.religion,
            hometown=req.hometown,
            residential_address=req.residential_address,
            nhis_number=req.nhis_number,
            ghana_card_number=req.ghana_card_number,
            is_boarding=req.is_boarding,
            orphan_status=req.orphan_status,
            disability=req.disability,
            is_active=True,
        )
        try:
            async with db.begin_nested():
                db.add(student)
                await db.flush()
            _log_row(db, batch.id, school_id, row_num, raw, "success", None, student.id)
            results.append(ImportRowResult(row=row_num, ref=req.admission_number, status="created", error=None))
            created += 1
        except IntegrityError:
            try:
                db.expunge(student)
            except Exception:
                pass
            msg = f"Admission number '{req.admission_number}' already exists at this school."
            _log_row(db, batch.id, school_id, row_num, raw, "error", msg)
            results.append(ImportRowResult(row=row_num, ref=req.admission_number, status="failed", error=msg))
            failed += 1

    batch.total_rows = created + failed
    batch.processed_rows = created + failed
    batch.error_count = failed
    batch.status = (
        ImportStatus.COMPLETED if failed == 0
        else ImportStatus.PARTIAL if created > 0
        else ImportStatus.FAILED
    )
    batch.completed_at = _utcnow()
    await db.flush()

    return ImportBatchResult(
        batch_id=batch.id,
        total_rows=created + failed,
        created=created,
        failed=failed,
        errors=[r for r in results if r.status == "failed"],
    )


def _log_row(
    db: AsyncSession,
    batch_id: uuid.UUID,
    school_id: uuid.UUID,
    row_num: int,
    raw: dict,
    status: str,
    error_message: str | None,
    entity_id: uuid.UUID | None = None,
) -> None:
    db.add(ImportRow(
        school_id=school_id,
        batch_id=batch_id,
        row_number=row_num,
        raw_data={k: str(v) if v is not None else None for k, v in raw.items()},
        status=status,
        error_message=error_message,
        entity_id=entity_id,
    ))
