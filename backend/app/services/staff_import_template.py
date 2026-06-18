"""
Branded Excel template generation for staff bulk import.

Called by:  GET /staff/import/template
Returns:    .xlsx bytes ready to stream to the browser

The template contains:
  - Row 1: Merged title with school name and brand_color
  - Row 2: Subtitle / instructions summary
  - Row 3: Column headers (locked)
  - Row 4: Sample/example row (greyed, italic)
  - Rows 5-204: Pre-formatted data-entry area with dropdowns for enum fields
  - Sentinel cell {SENTINEL_COL}1 = "TTEK_STAFF_IMPORT_{school_code}" (validated by importer)
  - Hidden "_positions" sheet (dropdown source)
  - "Instructions" sheet with field guide
"""
from __future__ import annotations
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation

from app.services.staff_import_constants import (
    _COLS, _DATA_START, _NUM_ROWS, make_sentinel, SENTINEL_COL,
)


def _argb(hex_color: str) -> str:
    return "FF" + hex_color.lstrip("#").upper()


def build_template(school_name: str, school_code: str, brand_color: str, position_names: list[str]) -> bytes:
    """Return branded .xlsx bytes for the staff import template."""
    wb = Workbook()

    _add_positions_sheet(wb, position_names)

    ws = wb.active
    ws.title = "Staff Data"

    brand_fill  = PatternFill("solid", fgColor=_argb(brand_color))
    grey_fill   = PatternFill("solid", fgColor="FFD9D9D9")
    sample_fill = PatternFill("solid", fgColor="FFF0F0F0")
    stripe_fill = PatternFill("solid", fgColor="FFE8EEF8")
    white_fill  = PatternFill("solid", fgColor="FFFFFFFF")
    hdr_font    = Font(bold=True, color="FFFFFFFF", name="Calibri", size=10)
    sample_font = Font(italic=True, color="FF888888", name="Calibri", size=10)
    body_font   = Font(name="Calibri", size=10)
    last_col    = _COLS[-1][0]

    # Row 1 — title (merge only covers data columns; sentinel sits to the right)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = f"{school_name}  —  Staff Import Template"
    c.font  = Font(bold=True, color="FFFFFFFF", name="Calibri", size=14)
    c.fill  = brand_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    ws[f"{SENTINEL_COL}1"].value = make_sentinel(school_code)

    # Row 2 — subtitle
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = (
        "Enter staff details from row 5 onward.  Required fields are marked *.  "
        "You may delete or overwrite the example row (row 4) before uploading."
    )
    c.font  = Font(italic=True, color="FF555555", name="Calibri", size=9)
    c.fill  = grey_fill
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    # Row 3 — column headers
    for col, label, _, _, width in _COLS:
        c = ws[f"{col}{3}"]
        c.value = label
        c.font  = hdr_font
        c.fill  = brand_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col].width = width
    ws.row_dimensions[3].height = 28

    # Row 4 — sample row
    sample_vals = [
        "TST001", "Kwame", "Adu", "Mensah", "MALE", "1985-06-15",
        "GHA-001234567", "C0012345678", "0244000000", "k.mensah@school.edu.gh",
        position_names[0] if position_names else "",
        "TEACHING", "PERMANENT", "MARRIED", "Accra, Greater Accra", "2020-09-01",
    ]
    for (col, *_), val in zip(_COLS, sample_vals):
        c = ws[f"{col}{4}"]
        c.value = val
        c.font  = sample_font
        c.fill  = sample_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 18

    # Rows 5+ — data area
    for row in range(_DATA_START, _DATA_START + _NUM_ROWS):
        fill = white_fill if row % 2 == 0 else stripe_fill
        for col, *_ in _COLS:
            c = ws[f"{col}{row}"]
            c.fill = fill
            c.font = body_font
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.protection = Protection(locked=False)
        ws.row_dimensions[row].height = 17

    # Dropdowns
    data_range = f"{_DATA_START}:{_DATA_START + _NUM_ROWS - 1}"

    dv_gender = DataValidation(type="list", formula1='"MALE,FEMALE"', allow_blank=True)
    dv_gender.sqref = f"E{_DATA_START}:E{_DATA_START + _NUM_ROWS - 1}"
    ws.add_data_validation(dv_gender)

    if position_names:
        dv_pos = DataValidation(
            type="list",
            formula1=f"'_positions'!$A$1:$A${len(position_names)}",
            allow_blank=True,
        )
        dv_pos.sqref = f"K{_DATA_START}:K{_DATA_START + _NUM_ROWS - 1}"
        ws.add_data_validation(dv_pos)

    dv_category = DataValidation(
        type="list", formula1='"TEACHING,NON_TEACHING"', allow_blank=True,
    )
    dv_category.sqref = f"L{_DATA_START}:L{_DATA_START + _NUM_ROWS - 1}"
    ws.add_data_validation(dv_category)

    dv_emptype = DataValidation(
        type="list",
        formula1='"PERMANENT,CONTRACT,NATIONAL_SERVICE,INTERN"',
        allow_blank=True,
    )
    dv_emptype.sqref = f"M{_DATA_START}:M{_DATA_START + _NUM_ROWS - 1}"
    ws.add_data_validation(dv_emptype)

    dv_marital = DataValidation(
        type="list",
        formula1='"SINGLE,MARRIED,DIVORCED,WIDOWED,SEPARATED"',
        allow_blank=True,
    )
    dv_marital.sqref = f"N{_DATA_START}:N{_DATA_START + _NUM_ROWS - 1}"
    ws.add_data_validation(dv_marital)

    ws.freeze_panes = f"A{_DATA_START}"
    ws.protection.sheet = True
    ws.protection.password = ""

    _add_instructions_sheet(wb, brand_fill, hdr_font)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _add_positions_sheet(wb: Workbook, position_names: list[str]) -> None:
    ws = wb.create_sheet("_positions")
    ws.sheet_state = "hidden"
    for i, name in enumerate(position_names, 1):
        ws.cell(row=i, column=1, value=name)


def _add_instructions_sheet(wb: Workbook, brand_fill: PatternFill, hdr_font: Font) -> None:
    ws = wb.create_sheet("Instructions")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 62
    rows = [
        ("TTEK Staff Import — Field Guide", None, True),
        (None, None, False),
        ("Field", "Notes / Valid Values", True),
        ("Staff Number*",   "Unique staff ID for this school. Required.",                       False),
        ("First Name*",     "Required.",                                                        False),
        ("Middle Name",     "Optional.",                                                        False),
        ("Last Name*",      "Required.",                                                        False),
        ("Gender",          "Select MALE or FEMALE from the dropdown.",                         False),
        ("Date of Birth",   "Format: YYYY-MM-DD  (e.g. 1985-06-15)",                           False),
        ("National ID",     "Ghana Card or passport number.",                                   False),
        ("SSNIT Number",    "SSNIT contributor ID (starts with C).",                            False),
        ("Phone",           "Mobile number, e.g. 0244000000.",                                  False),
        ("Email",           "Work or personal email address.",                                  False),
        ("Position",        "Select from the dropdown. Can be assigned later.",                 False),
        ("Staff Category",  "Select TEACHING or NON_TEACHING from the dropdown.",              False),
        ("Employment Type", "Select PERMANENT, CONTRACT, NATIONAL_SERVICE, or INTERN.",        False),
        ("Marital Status",  "Select from: SINGLE, MARRIED, DIVORCED, WIDOWED, SEPARATED.",    False),
        ("Address",         "Residential address.",                                             False),
        ("Date Joined",     "Date joined this school. Format: YYYY-MM-DD",                     False),
        (None, None, False),
        ("Tips", "", True),
        ("", "• Create all staff positions before downloading this template so the Position dropdown is complete.",  False),
        ("", "• Do not edit or delete header rows 1–3.",                                        False),
        ("", "• You may delete or overwrite the example row (row 4).",                          False),
        ("", "• Leave a cell blank if the information is not yet known.",                       False),
        ("", "• Save the file as .xlsx before uploading.",                                      False),
        ("", "• Duplicate staff numbers within the same school are rejected.",                  False),
    ]
    for i, (a, b, is_hdr) in enumerate(rows, 1):
        ca = ws.cell(row=i, column=1, value=a)
        cb = ws.cell(row=i, column=2, value=b)
        if is_hdr and a:
            ca.font = cb.font = hdr_font
            ca.fill = cb.fill = brand_fill
            if i == 1:
                ws.merge_cells("A1:B1")
                ws.row_dimensions[1].height = 24
